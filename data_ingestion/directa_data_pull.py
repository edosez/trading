# import keyring, 
import os, sys, glob
from pathlib import Path
from utils.utils import MyLogger, round_nearest_base, is_venv, load_config
import rpa as r
import pandas as pd
import numpy as np
import time
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import locale
from locale import atof

class DirectaDataPull:

    '''
    ## Create a bot that automatically logins to https://www1.directatrading.com/dlogin/PdL3v14159/ using the credential provided in keyring file with SPACENAME as directa
    '''

    def __init__(self, save_log = True, expiration_date_of_interest = 'GIU22', symbol = 'FXM22'):
        '''
        Constructor method
        '''
        self.save_log = save_log
        self.expiration_date_of_interest = expiration_date_of_interest
        self.symbol = symbol

        # Setting the yearmonth text for BarChart.com query
        locale.setlocale(locale.LC_ALL, 'it_IT.UTF-8')
        self.yearmonth_dt = pd.to_datetime(expiration_date_of_interest, format = '%b%y')
        locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
        self.yearmonth_eng = self.yearmonth_dt.strftime('%b %Y')

        # Defining the absolute path where the pandas profiling configuration files reside
        # the working directory should be offline_modelling\example\models\yyyy_mm_dd_hh_mm_ss and so it needs 3 jumps on top to get to offline_modelling\
        if os.path.basename(os.getcwd()) == 'trading':
            self.parent_path = os.getcwd()
        else:
            self.parent_path = str(Path().resolve().parents[2])

        if save_log:
            # Initiate the logging
            self._logging = MyLogger(log_file='logs/directa_data_pull.log', name='directa_data_pull')
        elif not save_log:
            self._logging = MyLogger(log_file=None, name='directa_data_pull')
        else:
            sys.exit('save_log parameter has not been set correctly | Adjust accordingly to either True or False')

        # Checking whether the script is running in a virtual environment
        if is_venv():
            self._logging.info('inside virtualenv or venv')
        else:
            self._logging.info('outside virtualenv or venv')

        self._logging.info("Executable is {}".format(sys.executable))

    def downloading_market_prices(self, session):

        '''
        Function that reads data from "Tabellone" and store data in a csv file 

        :param session: RPA session object
        '''

        session.click('#menugif')
        session.click('//*[@id="MenuOpt"]')
        self._logging.info("Going to the option ruler")

        current_expiration_string = session.read('//*[@id="wlbody"]/div[1]/table/thead/tr[1]/th/div[1]')
        self._logging.info("String for the selected expiration date is {}".format(current_expiration_string))
        self.current_exp_date = current_expiration_string.split('\n')[0][-5:]
        self._logging.info("String for the selected expiration date is {}".format(self.current_exp_date))


        if self.current_exp_date == self.expiration_date_of_interest:
            # In case desidred expiration date is already selected, then no need to click on the dropdown menù
            self._logging.info("The desired expiration date is already selected")
        else:
            # In case the desired expiration date is not selected, then click on the dropdown menù to change period
            self._logging.info("Changing the period of interest")
            self._logging.info("Clicking the dropdown menù for expiration date")
            r.click('//*[@id="wlbody"]/div[1]/table/thead/tr[1]/th/div[1]/i')
            # First button in the dropdown menù, containing options expiration dates
            first_button = session.read('/html/body/div[14]/div[1]/div[1]/table/thead/tr[1]/th/div[2]/button[1]')[-5:]
            self._logging.info("First expiration date in the list is {}".format(first_button))
            # Second button in the dropdown menù, containing options expiration dates
            second_button = session.read('/html/body/div[14]/div[1]/div[1]/table/thead/tr[1]/th/div[2]/button[2]')[-5:]
            self._logging.info("Second expiration date in the list is {}".format(second_button))

            # Dict containing the two expiration dates available on Directa
            dict_buttons = {first_button: '/html/body/div[14]/div[1]/div[1]/table/thead/tr[1]/th/div[2]/button[1]', 
                    second_button: '/html/body/div[14]/div[1]/div[1]/table/thead/tr[1]/th/div[2]/button[2]'
                    }

            # Iterating over the expiration dates dict and click the button that corresponds to the desired expiration date as a parameter
            for key, value in dict_buttons.items():
                if key == self.expiration_date_of_interest:
                    self._logging.info("Selecting the button for expiration date {}".format(key))
                    session.click(value)
    
            # reading the new expiration date so it can be stored on the MariaDB
            current_expiration_string = r.read('//*[@id="wlbody"]/div[1]/table/thead/tr[1]/th/div[1]')
            self._logging.info("String for the selected expiration date is {}".format(current_expiration_string))
            self.current_exp_date = current_expiration_string.split()[2]
            self._logging.info("Selected expiration date is {}".format(self.current_exp_date))
        
        self._logging.info("Downloading options data")
        # Reading the table containing the data
        session.table('//*[@id="wlbody"]/div[1]/table', "data\options_table.csv")
        self._logging.info("Options data download is completed")

    def downloading_open_positions(self, session):

        '''
        Function that reads data from "option ruler" and store data in a csv file 
        
        :param session: RPA session object
        '''
        self._logging.info("Downloading 'tabellone'")
        session.table('//*[@id="tab"]', "data/tabellone.csv")
        self._logging.info("'Tabellone' download is completed")

    def downloading_calendar_prices(self, session):
        '''
        Function that downloads table //*[@id="wlbody"]/div[8]/div/table at xpath //*[@id="wlbody"]/div[1]/table/thead/tr[1]/td[5]/i
        '''
        self._logging.info("Downloading 'calendario'")
        session.click('//*[@id="wlbody"]/div[1]/table/thead/tr[1]/td[5]')
        session.table('//*[@id="wlbody"]/div[8]/div/table', "data/options_calendar.csv")
        self._logging.info("'Calendario' download is completed")

    def downloading_greeks(self, session):
        '''
        Function that downloads table from BarChart.com website//*[@id="wlbody"]/div[8]/div/table at xpath //*[@id="wlbody"]/div[1]/table/thead/tr[1]/td[5]/i
        '''

        cred = load_config("cred")
        barchart_user = cred.get("BarChart.com").get("user")
        barchart_pwd = cred.get("BarChart.com").get("password")

        r.url('https://www.barchart.com/eu')
        time.sleep(10)

        # Cookie acceptance
        session.frame()
        if session.present('/html/body/div[9]/div[1]/div[1]/div/button[1]'):
            self._logging.info("Cookie rejection")
            session.click('/html/body/div[9]/div[1]/div[1]/div/button[1]')
            time.sleep(5)
        else:
            self._logging.info('Cookie already cleared')

        # Login
        session.click('//*[@id="bc-main-content-wrapper"]/div/div[1]/div[1]/div/div/div[2]/div[1]/a[1]')
        if session.present('//*[@id="bc-login-form"]/div[1]/input'):
            self._logging.info("Logging in to BarChart.com")
            time.sleep(2)
            session.type('//*[@id="bc-login-form"]/div[1]/input', "[clear]")
            session.type('//*[@id="bc-login-form"]/div[1]/input', "{}".format(barchart_user))
            time.sleep(2)
            session.type('//*[@id="login-form-password"]', "[clear]")
            session.type('//*[@id="login-form-password"]', "{}".format(barchart_pwd))
            session.click('//*[@id="bc-login-form"]/div[4]/button')
        else:
            self._logging.info("Already logged in to BarChart.com")

        # Searching for symbol page
        time.sleep(5)
        self._logging.info("Searching for {}".format(self.symbol))
        session.type('//*[@id="search"]', "{}[enter]".format(self.symbol))
        session.click('//*[@id="bc-main-content-wrapper"]/div/div[2]/div[1]/div/div[2]/div[2]/div/ul/li[5]/ul/li[2]/a')
        # Month of interest
        session.select('//*[@id="bc-options-toolbar__dropdown-month"]', self.yearmonth_eng)
        # Stacked data
        session.select('//*[@id="main-content-column"]/div/div[3]/div/div[3]/select', 'Stacked')
        # All strikes
        session.select('//*[@id="main-content-column"]/div/div[3]/div/div[2]/select', 'Show All') 
        time.sleep(5)
        self._logging.info("Downloading 'greeks'")
        session.click('//*[@id="main-content-column"]/div/div[3]/div/div[4]/a')
        time.sleep(5)
        self._logging.info("'Greeks' download is completed")
        try:
            self.newest_file = max(glob.iglob('/home/edoardo-sezzi/Desktop/trading/*.csv'), key=os.path.getctime).split('/')[-1]   
            self._logging.info("Newest file is {}".format(self.newest_file))   
            os.rename("/home/edoardo-sezzi/Desktop/trading/{}".format(self.newest_file), "data/greeks/{}".format(self.newest_file))
            self._logging.info("csv file has been moved to data folder")
        except ValueError:
            self._logging.warning("File is not placed in the directory, checking what's the latest file in data/greeks directory")
            self.newest_file = max(glob.iglob('/home/edoardo-sezzi/Desktop/trading/data/greeks/*.csv'), key=os.path.getctime).split('/')[-1]    

    def navigating_directa(self, options_prices = True, options_calendar = True, options_open_positions = True, options_greeks = True):
        '''
        Function that navigate the directa website and download the data through Option Ruler. 
        It allows to navigate the data based on desired expiration date
        
        :param expiration_date_of_interest: expiration date of interest, in form of MMMYY, e.g. GIU22 - remember to use the correct abbreviation for the month, in italian
        :type: str
        :return None       
        '''

        # NAMESPACE = "directa"
        # ENTRY = "B5665"
        # cred = keyring.get_credential(NAMESPACE, ENTRY)
        cred = load_config("cred")
        directa_user = cred.get("directa").get("user")
        directa_pwd = cred.get("directa").get("password")

        # Downloading the data from the website and store it into csv file
        self._logging.info("Initializing the session")
        r.init()
        r.url('https://www1.directatrading.com/dlogin/PdL3v14159/')
        self._logging.info("Connecting to the website")
        r.type('//*[@id="USER"]', "{}[enter]".format(directa_user))
        r.type('//*[@id="PASSW"]', "{}[enter]".format(directa_pwd))
        r.click('/html/body/div[1]/div[2]/div[1]/form/div[5]/div[1]/div[3]/button')
        self._logging.info("Login successful")
        time.sleep(15)
        self._logging.info("Selecting the frame so it is possible to navigate the personal page")
        r.frame("dliteframe")
        if r.present('//*[@id="bottoneavanti"]'):
            r.click('//*[@id="bottoneavanti"]')
        else:
            self._logging.info("No screen page with 'bottone avanti'")

        # Downloading "Tabellone" data
        if options_open_positions:
            self.downloading_open_positions(r)

        # Downloading "Option Ruler" data
        if options_prices:
            self.downloading_market_prices(r)

        # Downloading "Calendario" data
        if options_calendar:
            self.downloading_calendar_prices(r)

        # Downloading Greeks
        if options_greeks:
            self.downloading_greeks(r)

        r.close()
        self._logging.info("Closing the web session")

    def cleaning_options_data(self, csv_for_date = 'options_table.csv'):
        '''
        Function that load a csv file downloaded from Directa website and clean it. It adds few variables to allow a smooth loading on a RDBMS DB.
        
        :param csv_for_date: name of csv file downloaded from Directa website. Default is 'options_table.csv'. the function extract the creation time of the file so it will use that for the composition of pk
        :type: str
        :return df: dataframe containing the cleaned data that can be loaded onto the DB
        :rtype: pandas.DataFrame       
        '''

        # read csv file options_table.csv and extract the data
        cols = ['call_delta', 'call_volume', 'call_bid', 'call_median_price', 'call_ask', 'call_open_interest', 'call_price', 'strike',
                'put_price', 'put_volume', 'put_bid', 'put_median_price', 'put_ask', 'put_open_interest', 'put_delta']
 
        self._logging.info("Reading csv file for options data")     
        df = pd.read_csv(
            "data/options_table.csv", skiprows=5, names=cols, 
            thousands='.', decimal=',', usecols = np.arange(1, 16), na_values=['·', 'close']
            )
        df_call, df_put = df.iloc[: , :8], df.iloc[: , 7:]
        df_call.loc[:, 'option_type'] = 'C'
        df_put.loc[:, 'option_type'] = 'P'
        call_cols = [s.replace('call_', '') for s in df_call.columns]
        put_cols = [s.replace('put_', '') for s in df_put.columns] 
        df_call.columns, df_put.columns = call_cols, put_cols
        self._logging.info("Formatting DataFrame from wide to long")   
        df_long = pd.concat([df_call, df_put], ignore_index=False)
        self.insert_date = datetime.fromtimestamp(os.path.getctime('data/{0}'.format(csv_for_date))).strftime('%Y-%m-%d')
        df_long['insert_date'] = self.insert_date
        df_long['update_time'] = datetime.now()
        df_long['expiration_date'] = self.current_exp_date
        sql_pk = ["strike", "insert_date", "expiration_date", "option_type"]
        # df_long['pk'] = df_long.loc[:, ['strike', 'insert_date', 'expiration_date', 'option_type']].astype(str).agg('-'.join, axis = 1)
        future = pd.read_csv('data/options_table.csv', skiprows=0, nrows=1, usecols = [3], thousands='.', decimal=',').columns[0]
        self.future = float(future[:future.rfind('\n')].replace('.','').replace(',','.'))
        df_long.to_excel("data/options_table_clean.xlsx", index=False)
        df_long.to_csv("data/options_table_clean.csv", index=False)
        self._logging.info("Data cleaning for options data is completed and csv/xlsx files have been generated and saved")

        return df_long, sql_pk

    def cleaning_tabellone_data(self, purchase_date = None):
        '''
        :param purchase_date: date of purchase of the option. Default is None. format should be 'yyyy-mm-dd'
        :type: str
        '''

        locale.setlocale(locale.LC_NUMERIC, '')

        self._logging.info("Adding information to 'tabellone")
        cols_traded = ['symbol', 'description', 'current_price', 'benchmark', 'trend_perc', 'qty', 'price',
                        'gain_loss_abs', 'gain_loss_perc', 'recovery']
        df = pd.read_csv(
            "data/tabellone.csv", doublequote=False, skiprows = 1, names=cols_traded, 
            thousands='.', decimal=',', usecols = np.arange(0, 10), na_values=['·', 'close']
            ).iloc[:-1, :]
        cols_to_check = ['trend_perc', 'price', 'gain_loss_abs', 'gain_loss_perc', 'recovery']
        df[cols_to_check] = df[cols_to_check].replace({'\+': '', '%': '', '€': ''}, regex=True)
        df['purchase_date'] = pd.to_datetime(purchase_date, format='%Y-%m-%d')
        df['underlying_asset'] = df['description'].str.split('[" ".]').str[1]
        df['expiration_date'] = pd.to_datetime(df['description'].str.split().str[3].str[:2] + '-' + df['description'].str.split().str[3].str[2:] , format='%y-%m')
        df['option_type'] = df['description'].str.split().str[1]
        df['strike'] = df['description'].str.split().str[2]
        df['update_time'] = datetime.now()
        sql_pk = ["strike", "purchase_date", "expiration_date", "option_type", "underlying_asset"]
        # df['pk'] = df.loc[:, ['description', 'purchase_date']].astype(str).agg('-'.join, axis = 1)
        # Changing decimals from comma to dot and viceversa for thousands
        df[cols_to_check] = df[cols_to_check].applymap(atof)
        # dividing percentage columns by 100
        for col in df.columns[df.columns.str.contains('perc')]:
            df[col] = df[col]/100
    
        return df, sql_pk

    def cleaning_greeks_data(self):
        '''
        Function that imports csv file located in data/greeks folder, remove the column "Symbol" and store a pandas DataFrame ready to be load on DB.
        '''
            
        self._logging.info("Reading csv file for greeks data")     
        headers = [
            'strike', 'option_type', 'last','IV','delta',
            'gamma','theta','vega','IV_skew', 'insert_date'
            ]
        df = pd.read_csv("data/greeks/{}".format(self.newest_file), skipfooter=1, engine='python', header = 0, names = headers)
        df.loc[df['option_type'] == 'Call', 'option_type'] = 'C'
        df.loc[df['option_type'] == 'Put', 'option_type'] = 'P'
        df.loc[df['insert_date'].str.contains("CT"), 'insert_date'] = date.today().strftime('%m/%d/%y')
        df['insert_date'] = pd.to_datetime(df['insert_date'], format='%m/%d/%y')
        cols_to_check = ['IV', 'IV_skew']
        df[cols_to_check] = df[cols_to_check].replace({'\+': '', '%': '', '€': ''}, regex=True).astype(float)
        df['expiration_date'] = self.yearmonth_dt
        df['update_time'] = datetime.now()
        sql_pk = ["strike", "option_type", "expiration_date", "insert_date"]
        self._logging.info("Data cleaning for greeks data is completed and ready to be loaded on DB")

        return df, sql_pk


    def cleaning_calendar_data(self):
        '''
        Function that load a csv file downloaded from Directa website and clean it. It adds few variables to allow a smooth loading on a RDBMS DB.
        
        :param csv_for_date: name of csv file downloaded from Directa website. Default is 'options_table.csv'. the function extract the creation time of the file so it will use that for the composition of pk
        :type: str
        :param purchase_date: date of purchase of the option. Default is None. format should be 'yyyy-mm-dd'
        :type: str
        :return df: dataframe containing the cleaned data that can be loaded onto the DB
        :rtype: pandas.DataFrame       
        '''

        self._logging.info("Importing calendar data file and cleaning it") 
        df = pd.read_csv('data/options_calendar.csv', skiprows=1, na_values=['·', 'close'], thousands='.', decimal=',')
        df.columns = df.columns.str.lower()
        self._logging.info("Melting from wide to long")
        df_long = df.melt(id_vars='strike', var_name = 'expiration_date', value_name = 'price')
        df_long.loc[:, 'option_type'] = ['P' if x == '.1' else 'C' for x in df_long['expiration_date'].str[-2:]]
        df_long.loc[:, 'expiration_date'] = df_long.loc[:, 'expiration_date'].str.replace('.1', '', regex = False)
        df_long.loc[:, 'expiration_date'] = pd.to_datetime(df_long.loc[:, 'expiration_date'], format='%d-%m-%Y')
        sql_pk = ["strike" , "expiration_date", "option_type"]
        # df_long['pk'] = df_long.loc[:, ['strike', 'expiration_date', 'option_type']].astype(str).agg('-'.join, axis = 1)
        df_long['insert_date'] = date.today().strftime('%Y-%m-%d')
        df_long.to_csv('data/options_calendar_clean.csv', index=False)
        self._logging.info("Data cleaning for calendar data is completed and csv/xlsx files have been generated and saved")

        return df_long, sql_pk


    def editing_strategy_calculator(self, input_df, grid_shift_input = 25):
        '''
        Function that iterates through row of spreadsheet and, depending on the type of option,
        populate the ceels with options prices.

        :param df: pandas DataFrame containing the data to be populated in strategy calculator (put/call options prices)
        :type: pandas.DataFrame
        :return: None
        '''

        self._logging.info("Editing strategy calculator")

        # Editing strategy calculator

        grid_shift = grid_shift_input
        first_row_excel = 13
        last_row_excel = 43
        option_fee = 2.5

        # Dict defining call and put settings for writing the strategy calculator
        dict_strat_cal_cells = {'put': [6, 'P'], 'call': [2, 'C']}

        self._logging.info("Opening the strategy calculator")
        workbook_strategy_calculator = load_workbook('strategy_calculator/STRATEGY.xlsx')
        worksheet = workbook_strategy_calculator['EUROSTOXX50']

        # Insert current future value
        worksheet['D5'] = self.future
        worksheet['D28'] = round_nearest_base(self.future, base = grid_shift)
        worksheet['F2'] = grid_shift
        worksheet['E49'] = option_fee

        # filter input_df using only strie divisible by grid_shift
        clean_df = input_df[input_df['strike'] % grid_shift == 0].reset_index()

        self._logging.info("Strikes in analysis are {}".format(clean_df['strike'].unique()))

        self._logging.info('Rounded future is {}'.format(round_nearest_base(self.future, base = grid_shift)))

        min_index = int(clean_df.median_price[clean_df.strike == round_nearest_base(self.future, base = grid_shift)].index[0] - (last_row_excel-first_row_excel)/2)
        max_index = int(clean_df.median_price[clean_df.strike == round_nearest_base(self.future, base = grid_shift)].index[0] + (last_row_excel-first_row_excel)/2) 

        self._logging.info('Minimum index is {}'.format(min_index))
        self._logging.info('Maximum index is {}'.format(max_index))

        for key, value in dict_strat_cal_cells.items():

            self._logging.info("Populating the strategy calculator with {} options".format(key))
            col_excel = value[0]
            prices = clean_df[clean_df['option_type'] == value[1]].iloc[min_index: max_index+1]['median_price'].to_list()

            self._logging.info("Writing the strategy calculator xlsx file iterating over the rows")
            for i, row in enumerate(worksheet.iter_rows(min_row=first_row_excel, min_col=col_excel, max_row=last_row_excel, max_col=col_excel)):    
                for cell in row:
                    self._logging.info("Price written is {}".format(prices[i]))
                    worksheet[get_column_letter(cell.column) + str(cell.row)] = prices[i]

        self._logging.info("Saving the strategy calculator xlsx file")
        workbook_strategy_calculator.save("strategy_calculator/Strategy_{0}_{1}.xlsx".format(self.insert_date, self.current_exp_date))

        self._logging.info("Editing strategy calculator is completed and file strategy_calculator/Strategy_{0}_{1}.xlsx has been saved".format(self.insert_date, self.current_exp_date))





